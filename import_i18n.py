import os
import polib
from openpyxl import load_workbook


def update_translations():
    """更新各语言PO文件，若不存在则自动创建"""
    # 检查Excel文件是否存在
    excel_file = "i18n.xlsx"
    if not os.path.exists(excel_file):
        print(f"错误：{excel_file} 不存在，请检查文件是否存在")
        return

    wb = load_workbook(excel_file)
    ws = wb.active

    # 读取语言列表（从第一行获取）
    languages = [cell.value for cell in ws[1] if cell.value]  # 过滤空值

    if not languages:
        print("错误：未从Excel第一行读取到任何语言信息")
        return

    # 构建翻译数据 {msgid: {lang: translation}}
    translations = {}
    for row in ws.iter_rows(min_row=2):
        msgid = row[0].value
        if not msgid:
            continue  # 跳过空的msgid行

        translations[msgid] = {
            lang: row[i].value or "" for i, lang in enumerate(languages)
        }

    # 更新各语言文件
    for lang in languages:
        lang_dir = os.path.join(".", lang)
        po_path = os.path.join(lang_dir, f"{lang}.po")

        # 检查并创建语言目录
        if not os.path.exists(lang_dir):
            print(f"目录 {lang_dir} 不存在，正在创建...")
            os.makedirs(lang_dir, exist_ok=True)  # exist_ok=True 避免目录已存在时出错

        # 检查并创建PO文件
        if not os.path.exists(po_path):
            print(f"文件 {po_path} 不存在，正在创建新文件...")
            # 创建新的PO文件并设置基本信息
            po = polib.POFile()
            po.metadata = {
                "Project-Id-Version": "1.0",
                "Report-Msgid-Bugs-To": "",
                "POT-Creation-Date": "",
                "PO-Revision-Date": "",
                "Last-Translator": "",
                "Language-Team": "",
                "MIME-Version": "1.0",
                "Content-Type": "text/plain; charset=utf-8",
                "Content-Transfer-Encoding": "8bit",
                "Language": lang,
            }
        else:
            # 加载已存在的PO文件
            po = polib.pofile(po_path)

        # 创建现有条目的映射，方便快速查找
        existing = {entry.msgid: entry for entry in po}

        # 更新现有条目并添加新条目
        for msgid, trans in translations.items():
            msgstr = trans.get(lang, "")

            if msgid in existing:
                # 更新现有条目的翻译
                existing[msgid].msgstr = msgstr
            else:
                # 添加新的翻译条目
                new_entry = polib.POEntry(msgid=msgid, msgstr=msgstr)
                po.append(new_entry)

        # 保存更新后的PO文件
        po.save(po_path)
        print(f"已成功更新 {po_path}")


if __name__ == "__main__":
    update_translations()
